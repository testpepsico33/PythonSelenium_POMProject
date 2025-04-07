import openpyxl


def get_row_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_row  # this will return number of row in a sheet


def get_column_count(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_column  ##this will return number of column in a sheet


def get_cell_data(path, sheet_name, row_number, column_number):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_number, column=column_number).value


def set_cell_data(path, sheet_name, row_number, column_number, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_number, column=column_number).value = data #writing data in a excel sheet
    workbook.save(path)

def get_data_from_excel(path, sheet_name):
    final_list=[]
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheet_name] # sheet name of given excel
    total_row =sheet.max_row
    total_colum =sheet.max_column

    for r in range(2,total_row+1):
        row_list=[]
        for c in range(1,total_colum+1):
            row_list.append(sheet.cell(row=r,column=c).value)
        final_list.append(row_list)
    return final_list
